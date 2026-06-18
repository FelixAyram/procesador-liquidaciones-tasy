"""Exportación y formato legible de Excel."""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from excel_schema import (
    AMOUNT_COLUMN_INDEXES,
    DATE_COLUMN_INDEXES,
    INTEGER_COLUMN_INDEXES,
    row_to_values,
)
from theme import COLORS


def _cell_display_length(value) -> int:
    if value is None:
        return 0
    if isinstance(value, float):
        if value == int(value):
            text = f"{int(value):,}".replace(",", ".")
        else:
            text = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return len(text)
    return len(str(value))


def format_excel_worksheet(worksheet) -> None:
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=COLORS["header_excel"].replace("#", ""))
    body_font = Font(name="Segoe UI", size=11, color="1E293B")
    zebra_fill = PatternFill("solid", fgColor="F8FAFC")

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.font = body_font

            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                if col_idx in AMOUNT_COLUMN_INDEXES:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx in DATE_COLUMN_INDEXES:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx in INTEGER_COLUMN_INDEXES:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 36

    for col_idx in range(1, max_col + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            max_length = max(max_length, _cell_display_length(cell.value))
        header_value = worksheet.cell(row=1, column=col_idx).value
        header_len = len(str(header_value)) if header_value is not None else 8
        width = max(max_length, header_len) + 2
        if col_idx in AMOUNT_COLUMN_INDEXES:
            width = max(width, 14)
        elif col_idx in {2, 24}:
            width = min(max(width, 20), 38)
        else:
            width = min(width, 30)
        worksheet.column_dimensions[column_letter].width = width

    worksheet.auto_filter.ref = worksheet.dimensions


def _next_append_row(worksheet) -> int:
    """Primera fila vacía después de los datos existentes (debajo del encabezado)."""
    if worksheet.max_row <= 1:
        return 2

    last_with_data = 1
    max_col = max(worksheet.max_column, 1)
    for row_idx in range(2, worksheet.max_row + 1):
        if any(
            worksheet.cell(row=row_idx, column=col).value not in (None, "")
            for col in range(1, max_col + 1)
        ):
            last_with_data = row_idx
    return last_with_data + 1


def write_rows_to_excel(template_path: Path, output_path: Path, rows: list[dict]) -> None:
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    start_row = _next_append_row(worksheet)
    for offset, row_data in enumerate(rows):
        values = row_to_values(row_data)
        for col_idx, value in enumerate(values, start=1):
            worksheet.cell(row=start_row + offset, column=col_idx, value=value)

    format_excel_worksheet(worksheet)
    workbook.save(output_path)
    workbook.close()
